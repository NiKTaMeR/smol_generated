import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from datetime import datetime
from data_ingestion import ingest_data

def parse_information():
    data = ingest_data()
    earliest_date = data['Transaction Date'].min()
    latest_date = data['Transaction Date'].max()

    wb = Workbook()
    ws = wb.active
    ws.title = "Inputs_Variables"

    ws['A1'] = "threshold_contraction"
    ws['B1'] = 0.1  # Default percentage value, user can modify it later
    ws['A2'] = "threshold_upsell"
    ws['B2'] = 0.1  # Default percentage value, user can modify it later

    output_filename = input("Enter the path and filename for the new Excel file: ")
    wb.save(output_filename)

    print("Now opening the excel file, fill in the monthly results, save and close the file, before coming back here.")
    os.startfile(output_filename)

    input("Press Enter after you have filled in the monthly results and saved the file: ")

    return output_filename, earliest_date, latest_date