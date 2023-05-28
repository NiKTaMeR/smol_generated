import os
import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

def get_file_path(file_type):
    file_path = input(f"Enter the path of the {file_type} file: ")
    while not os.path.exists(file_path):
        print("File not found. Please try again.")
        file_path = input(f"Enter the path of the {file_type} file: ")
    return file_path

def get_month_range(data):
    min_date = min(data, key=lambda x: x["transaction_date"])["transaction_date"]
    max_date = max(data, key=lambda x: x["transaction_date"])["transaction_date"]
    min_month = datetime.date(min_date.year, min_date.month, 1)
    max_month = datetime.date(max_date.year, max_date.month, 1)
    return min_month, max_month

def create_monthly_headers(min_month, max_month):
    headers = []
    current_month = min_month
    while current_month <= max_month:
        headers.append(current_month.strftime("%m/%Y"))
        current_month = current_month + datetime.timedelta(days=32)
        current_month = datetime.date(current_month.year, current_month.month, 1)
    return headers

def create_excel_file(file_path):
    wb = Workbook()
    wb.save(file_path)
    return wb

def open_excel_file(file_path):
    return load_workbook(file_path)

def save_and_close_excel_file(wb, file_path):
    wb.save(file_path)
    wb.close()

def create_worksheet(wb, sheet_name):
    if sheet_name in wb:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(sheet_name)
    return ws

def write_headers(ws, headers, start_row=1, start_col=1):
    for col, header in enumerate(headers, start=start_col):
        ws.cell(row=start_row, column=col, value=header)

def write_data(ws, data, start_row=2, start_col=1):
    for row, row_data in enumerate(data, start=start_row):
        for col, value in enumerate(row_data, start=start_col):
            ws.cell(row=row, column=col, value=value)

def get_column_index(ws, column_name):
    for col in range(1, ws.max_column + 1):
        if ws.cell(row=1, column=col).value == column_name:
            return col
    return None

def get_row_index(ws, row_name, column_index=1):
    for row in range(2, ws.max_row + 1):
        if ws.cell(row=row, column=column_index).value == row_name:
            return row
    return None

def get_cell_value(ws, row, column):
    return ws.cell(row=row, column=column).value

def set_cell_value(ws, row, column, value):
    ws.cell(row=row, column=column, value=value)

def format_percentage(value):
    return f"{value:.2%}"